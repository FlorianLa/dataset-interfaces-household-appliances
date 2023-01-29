import cv2
import numpy as np
import os
import sys
import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
import xmltodict

def xmlToJson (file):

    with open(sys.argv[1]+"/"+file) as xml_file:
        data_dict = xmltodict.parse(xml_file.read())
        xml_file.close()

        # generate the object using json.dumps()
        # corresponding to json data

        json_data = json.dumps(data_dict)
        # Write the json data to output
        # json file
        with open("xml/" + file[:-3]+"json", "w") as json_file:
            json_file.write(json_data)
            json_file.close()

        calculateXMLJson(file[:-3]+"json")

def calculateXMLJson (file):
    with open("xml/"+file, "r") as read_file:
        data = json.load(read_file)
    wb = openpyxl.load_workbook("dataset.xlsx")
    sheet = wb.get_sheet_by_name('Sheet1')
   # for key in data['annotation'].keys():
   #     print(key)
   #     print(data['annotation'][key])
   #     print("--------------")
    objects = data['annotation']['object']
    maxRow = sheet.max_row
    row = maxRow + 1

    pushbutton = 0
    knob = 0
    toggle = 0
    slider = 0
    others = 0
    sizes = []
    pixels = []
    name = data['annotation']['filename']
    print(data['annotation'])
    print("-------------")
    try:
        print(objects['name'])
        label = objects['name']
        height = float(objects['bndbox']['ymax']) - float(objects['bndbox']['ymin'])
        width = float(objects['bndbox']['xmax']) - float(objects['bndbox']['xmin'])

        if label == 'pushbutton':
            pushbutton += 1
        elif label == 'knob':
            knob += 1
        elif label == 'slider':
            slider += 1
        elif label == 'toggle':
            toggle += 1
        else:
            others += 1
        sizes.append(width * height)
        pixels.append((width, height))

    except:
        print(objects[0])
        for i in range(0, len(objects)):

            label = objects[i]['name']
            height =float(objects[i]['bndbox']['ymax']) - float(objects[i]['bndbox']['ymin'])
            width = float(objects[i]['bndbox']['xmax']) - float(objects[i]['bndbox']['xmin'])

            if label == 'pushbutton':
                pushbutton += 1
            elif label == 'knob':
                knob += 1
            elif label == 'slider':
                slider += 1
            elif label == 'toggle':
                toggle += 1
            else:
                others += 1
            sizes.append(width * height)
            pixels.append((width, height))
        print("sizes: ", sizes)

    indexMin = sizes.index(min(sizes))
    indexMax = sizes.index(max(sizes))
    print(indexMin)
    minSize = str(pixels[indexMin][0]) + "x" + str(pixels[indexMin][1])
    print(minSize)
    maxSize = str(pixels[indexMax][0]) + "x" + str(pixels[indexMax][1])
    print(maxSize)

    # ID
    sheet.cell(row=row, column=1).value = name[:-4]
    # pushbuttons
    sheet.cell(row=row, column=2).value = pushbutton
    # knobs
    sheet.cell(row=row, column=3).value = knob
    # sliders
    sheet.cell(row=row, column=4).value = slider
    # toggles
    sheet.cell(row=row, column=5).value = toggle
    # totalElements
    sheet.cell(row=row, column=6).value = pushbutton + knob + slider + toggle + others
    # minSize
    sheet.cell(row=row, column=7).value = min(sizes)
    # minDimension
    sheet.cell(row=row, column=8).value = minSize
    # maxSize
    sheet.cell(row=row, column=9).value = max(sizes)
    # maxDimension
    sheet.cell(row=row, column=10).value = maxSize

    wb.save("dataset.xlsx")

def calculateJson (file):
    with open(sys.argv[1]+"/"+file, "r") as read_file:
        data = json.load(read_file)
    wb = openpyxl.load_workbook("dataset.xlsx")
    sheet = wb.get_sheet_by_name('Sheet1')
    wbOT = openpyxl.load_workbook("datasetOhneTouch.xlsx")
    sheetOT = wbOT.get_sheet_by_name('Sheet1')
   # maxCol = sheet.max_column
    maxRow = sheet.max_row
    row = maxRow + 1
   # print(maxCol)
   # print(maxRow)
    name = data[0]['image']
   # print(name[:-4])
   # print(data[0]['image'])
    #print("Hier")
   # print (sheet.cell(row = 1, column = 7).value)
    pushbutton = 0
    knob = 0
    touch = 0
    toggle = 0
    slider = 0
    others = 0
    sizes = []
    pixels = []

    # berechnen
    number = len(data[0]['annotations'])
    for i in range(0, number):
        label = data[0]['annotations'][i]['label']
        if label == 'pushbutton':
            pushbutton += 1
        elif label == 'knob':
            knob += 1
        elif label == 'touch':
            touch += 1
        elif label == 'slider':
            slider += 1
        elif label == 'toggle':
            toggle += 1
        else:
            others += 1
        cors = data[0]['annotations'][i]['coordinates']
        sizes.append(cors['width'] * cors['height'])
        pixels.append((cors['width'], cors['height']))
    print("sizes: ", sizes)


    indexMin = sizes.index(min(sizes))
    indexMax = sizes.index(max(sizes))
    print(indexMin)
    minSize = str(pixels[indexMin][0]) +"x"+str(pixels[indexMin][1])
    print(minSize)
    maxSize = str(pixels[indexMax][0]) +"x"+str(pixels[indexMax][1])
    print(maxSize)

    #befüllen
    valType = ""
    valLightning = ""
    valManufacturer = ""
    valModel = ""
    valVideo = ""
    valLength = ""
    valFormat = ""
    valNumber = ""

    for i in range(2, sheetOT.max_row):
        if int(sheetOT.cell(row=i, column=1).value[0:3]) == name[:-4]:
            if sheetOT.cell(row=i, column=11).value is not None:
                valType = sheetOT.cell(row=i, column=11).value
            if sheetOT.cell(row=i, column=12).value is not None:
                valLightning = sheetOT.cell(row=i, column=12).value
            if sheetOT.cell(row=i, column=13).value is not None:
                valManufacturer = sheet.cell(row=i, column=13).value
            if sheetOT.cell(row=i, column=14).value is not None:
                valModel = sheetOT.cell(row=i, column=14).value
            if sheetOT.cell(row=i, column=15).value is not None:
                valVideo = sheetOT.cell(row=i, column=15).value
            if sheetOT.cell(row=i, column=16).value is not None:
                valLength = sheetOT.cell(row=i, column=16).value
            if sheetOT.cell(row=i, column=17).value is not None:
                valFormat = sheetOT.cell(row=i, column=17).value
            if sheetOT.cell(row=i, column=18).value is not None:
                valNumber = sheetOT.cell(row=i, column=18).value
            break

    # ID
    sheet.cell(row=row, column=1).value = name[:-4]
    # pushbuttons
    sheet.cell(row=row, column=2).value = pushbutton
    # touch
    sheet.cell(row=row, column=3).value = touch
    #knobs
    sheet.cell(row=row, column=4).value = knob
    #sliders
    sheet.cell(row=row, column=5).value = slider
    #toggles
    sheet.cell(row=row, column=6).value = toggle
    #totalElements
    sheet.cell(row=row, column=7).value = pushbutton+knob+slider+toggle+others+touch
    #minSize
    sheet.cell(row=row, column=8).value = min(sizes)
    #minDimension
    sheet.cell(row=row, column=9).value = minSize
    #maxSize
    sheet.cell(row=row, column=10).value = max(sizes)
    #maxDimension
    sheet.cell(row=row, column=11).value = maxSize
    #type
    sheet.cell(row=row, column=12).value = valType
    #lightning
    sheet.cell(row=row, column=13).value = valLightning
    #manufacturer
    sheet.cell(row=row, column=14).value = valManufacturer
    #model
    sheet.cell(row=row, column=15).value = valModel
    #video
    sheet.cell(row=row, column=16).value = valVideo
    #videoLength
    if valLength != "":
        sheet.cell(row=row, column=17).value = int(valLength)
    #videoFormat
    sheet.cell(row=row, column=18).value = valFormat
    #numberOfImages
    if valNumber != "":
        sheet.cell(row=row, column=19).value = int(valNumber)

    wb.save("dataset.xlsx")
    #print(data[0]['annotations'][9]['coordinates']['x'])
    print("--------------------")
   # print(data)


for filename in os.listdir(sys.argv[1]):


    print("hier: ", filename)
    if filename.endswith(".JSON") or filename.endswith(".json"):
        calculateJson(filename)
        continue
    elif filename.endswith(".xml") or filename.endswith(".XML"):
        xmlToJson(filename)
        continue
    else:
        continue
